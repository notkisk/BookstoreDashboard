"""
Excel Export for Orders

This module preserves the exact formatting of the template Excel file
while adding order data rows.
"""

import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

def export_orders_to_excel(orders, output_path=None):
    """
    Export orders to Excel using the template file while preserving all formatting.
    
    Args:
        orders: List of order objects with customer and item details
        output_path: Path to save the Excel file. If None, a default path is used.
    
    Returns:
        Path to the saved Excel file
    """
    # Use the template file path
    template_path = 'templates/order_export_template.xlsx'
    
    # If no output path specified, create one with timestamp
    if not output_path:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_dir = 'exports'
        # Create exports directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f'orders_export_{timestamp}.xlsx')

    # Load the template workbook - with all its formatting, styles, etc.
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook.active
    
    # Determine the row to start adding data (usually row 2, after headers)
    start_row = 2
    
    # Get all the style attributes from the header row to match formatting
    header_font = sheet.cell(row=1, column=1).font
    header_fill = sheet.cell(row=1, column=1).fill
    header_border = sheet.cell(row=1, column=1).border
    header_alignment = sheet.cell(row=1, column=1).alignment
    
    # Add order data starting from row 2
    for i, order in enumerate(orders):
        row_num = start_row + i
        
        # Map order data to the columns
        # Column 1: reference commande
        sheet.cell(row=row_num, column=1).value = order.get('reference', '')
        
        # Column 2: nom et prenom du destinataire*
        if order.get('customer'):
            sheet.cell(row=row_num, column=2).value = order['customer'].get('name', '')
        
        # Column 3: telephone*
        if order.get('customer'):
            sheet.cell(row=row_num, column=3).value = order['customer'].get('phone', '')
        
        # Column 4: telephone 2
        if order.get('customer'):
            sheet.cell(row=row_num, column=4).value = order['customer'].get('phone2', '')
        
        # Column 5: code wilaya*
        if order.get('customer') and order['customer'].get('address'):
            sheet.cell(row=row_num, column=5).value = order['customer']['address'].get('wilayaId', '')
        
        # Column 6: wilaya de livraison
        if order.get('customer') and order['customer'].get('address'):
            sheet.cell(row=row_num, column=6).value = order['customer']['address'].get('wilayaName', '')
        
        # Column 7: commune de livraison*
        if order.get('customer') and order['customer'].get('address'):
            sheet.cell(row=row_num, column=7).value = order['customer']['address'].get('communeName', '')
        
        # Column 8: adresse de livraison*
        if order.get('customer') and order['customer'].get('address'):
            sheet.cell(row=row_num, column=8).value = order['customer']['address'].get('streetAddress', '')
        
        # Column 9: produit* (combine items)
        if order.get('items'):
            products = [f"{item['book']['title']} (x{item['quantity']})" for item in order['items']]
            sheet.cell(row=row_num, column=9).value = ", ".join(products)
        
        # Column 10: poids (kg)
        # Assuming weight is calculated or stored somewhere
        weight = 0.5  # Default weight (0.5 kg for books)
        if order.get('items'):
            # Calculate based on number of books, assuming 0.5kg per book
            weight = sum([item['quantity'] * 0.5 for item in order['items']])
        sheet.cell(row=row_num, column=10).value = weight
        
        # Column 11: montant du colis*
        sheet.cell(row=row_num, column=11).value = order.get('finalAmount', order.get('totalAmount', 0))
        
        # Column 12: remarque
        sheet.cell(row=row_num, column=12).value = order.get('notes', '')
        
        # Column 13-17: Special flags (default to empty)
        # FRAGILE, ECHANGE, PICK UP, RECOUVREMENT, STOP DESK
        for col in range(13, 18):
            sheet.cell(row=row_num, column=col).value = ""
            
        # Column 18: Lien map
        if order.get('customer') and order['customer'].get('address') and order['customer']['address'].get('mapLink'):
            sheet.cell(row=row_num, column=18).value = order['customer']['address'].get('mapLink', '')
        
        # Copy the cell styles from the header row to maintain consistent formatting
        for col in range(1, 19):  # 18 columns
            cell = sheet.cell(row=row_num, column=col)
            
            # Apply consistent styling but with white background
            cell.font = Font(
                name=header_font.name,
                size=header_font.size,
                bold=False  # Data rows typically not bold
            )
            
            # Use a lighter fill for data rows
            cell.fill = PatternFill(fill_type="solid", start_color="FFFFFF")
            
            # Keep the same border style
            cell.border = Border(
                left=Side(style=header_border.left.style) if header_border.left else None,
                right=Side(style=header_border.right.style) if header_border.right else None,
                top=Side(style=header_border.top.style) if header_border.top else None,
                bottom=Side(style=header_border.bottom.style) if header_border.bottom else None
            )
            
            # Keep the same alignment
            cell.alignment = Alignment(
                horizontal=header_alignment.horizontal,
                vertical=header_alignment.vertical,
                wrap_text=True  # Enable text wrapping
            )
    
    # Save the workbook to the output path
    workbook.save(output_path)
    return output_path