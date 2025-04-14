"""
EcoTrack Excel Export for Orders

This module preserves the EXACT formatting, formulas, and validation rules of the EcoTrack delivery template Excel file
while adding order data rows in the correct locations.

Special considerations:
1. Uses openpyxl with keep_vba=True to preserve all macros and special properties
2. Only adds data without modifying any existing cells containing formulas or validation rules
3. Handles proper data type conversions (dates, numbers, text, etc.)
4. Maps column headers intelligently for reliable exports
"""

import os
import json
from datetime import datetime
import logging
import openpyxl
from openpyxl.utils import get_column_letter
import argparse
import sys

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    filename='ecotrack_export.log'
)
logger = logging.getLogger('ecotrack_export')

# Constants
TEMPLATE_FILENAME = 'upload_ecotrack_v31.xlsx'
DEFAULT_TEMPLATE_PATH = os.path.join(os.getcwd(), 'templates', TEMPLATE_FILENAME)
ERROR_LOG_PATH = os.path.join(os.getcwd(), 'exports', 'ecotrack_export_errors.log')

class EcoTrackExcelExporter:
    """Handles exporting orders to the EcoTrack template Excel file."""
    
    def __init__(self, template_path=None):
        """
        Initialize the exporter with the template file path.
        
        Args:
            template_path: Path to the EcoTrack template file.
                           If None, uses the default path.
        """
        self.template_path = template_path or DEFAULT_TEMPLATE_PATH
        self.column_mapping = {}
        self.header_row = 1  # Default to first row if not detected
        self.data_start_row = 2  # Default to second row if not detected
        self.log_errors = []
        
    def export_orders(self, orders, output_path=None):
        """
        Export orders to an EcoTrack Excel file using their template.
        
        Args:
            orders: List of order objects with customer and item details
            output_path: Path to save the Excel file. If None, a default path is used.
        
        Returns:
            Path to the saved Excel file
        """
        try:
            # Verify template exists and is readable
            if not os.path.exists(self.template_path):
                error_msg = f"Template file not found at: {self.template_path}"
                logger.error(error_msg)
                raise FileNotFoundError(error_msg)
                
            # Try to read the template to verify it's a valid Excel file
            try:
                workbook = openpyxl.load_workbook(self.template_path, keep_vba=True, data_only=False)
            except Exception as e:
                error_msg = f"Failed to load template file - it may be corrupted or not a valid Excel file: {str(e)}"
                logger.error(error_msg)
                raise Exception(error_msg)

            # If no output path specified, create one with timestamp
            if not output_path:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_dir = os.path.join(os.getcwd(), 'exports')
                # Create exports directory if it doesn't exist
                os.makedirs(output_dir, exist_ok=True)
                output_path = os.path.join(output_dir, f'ecotrack_export_{timestamp}.xlsx')
            
            # Log the operation
            logger.info(f"Starting export of {len(orders)} orders to {output_path}")
            logger.info(f"Using template: {self.template_path}")
            
            # Get the active sheet (assumes the template uses the active sheet for data)
            sheet = workbook.active
            
            # Discover the column mappings and header row
            try:
                self._discover_column_mapping(sheet)
            except Exception as e:
                error_msg = f"Failed to discover column mappings: {str(e)}"
                logger.error(error_msg)
                raise Exception(error_msg)
            
            # Add order data
            try:
                self._add_order_data(sheet, orders)
            except Exception as e:
                error_msg = f"Failed to add order data: {str(e)}"
                logger.error(error_msg)
                raise Exception(error_msg)
            
            # Save the workbook to the output path
            try:
                workbook.save(output_path)
            except Exception as e:
                error_msg = f"Failed to save output file: {str(e)}"
                logger.error(error_msg)
                raise Exception(error_msg)
            
            # Log any errors
            if self.log_errors:
                self._write_error_log()
                
            logger.info(f"Excel export completed successfully: {output_path}")
            return output_path
            
        except Exception as e:
            error_msg = f"Error exporting orders to Excel: {str(e)}"
            logger.error(error_msg)
            self.log_errors.append(error_msg)
            self._write_error_log()
            raise
            
    def _discover_column_mapping(self, sheet):
        """
        Discover the column headers in the template to map our data correctly.
        Uses case-insensitive matching and handles synonyms.
        
        Args:
            sheet: The Excel worksheet object
        """
        # Header mapping with synonyms (lowercase for case-insensitive matching)
        header_synonyms = {
            'ref': ['ref', 'reference', 'reference commande', 'commande ref', 'réf', 'référence'],
            'name': ['nom', 'nom client', 'client', 'customer', 'customer name', 'destinataire', 'nom destinataire'],
            'phone': ['telephone', 'téléphone', 'tel', 'tél', 'phone', 'mobile', 'gsm', 'phone number', 'numéro'],
            'phone2': ['telephone 2', 'téléphone 2', 'telephone2', 'téléphone2', 'tel2', 'tél2', 'numéro téléphone2', 'numéro secondaire'],
            'wilaya_code': ['code wilaya', 'wilaya code', 'code', 'cod wilaya', 'code w', 'codew', 'code wilaya*'],
            'wilaya': ['wilaya', 'ville', 'city', 'province', 'region', 'région'],
            'commune': ['commune', 'commune de livraison', 'ville'],
            'address': ['adresse', 'address', 'adresse client', 'delivery address', 'adresse livraison', 'lieu'],
            'product': ['produit', 'product', 'article', 'désignation', 'designation'],
            'weight': ['poids', 'poids (kg)', 'weight', 'kg'],
            'amount': ['montant', 'montant du colis', 'prix', 'price', 'cod', 'c.o.d'],
            'note': ['remarque', 'note', 'notes', 'commentaire', 'observation'],
            'fragile': ['fragile'],
            'exchange': ['échange', 'echange', 'exchange'],
            'pickup': ['pick up', 'pickup', 'ramassage'],
            'cod': ['recouvrement', 'c.o.d', 'cod', 'contre remboursement'],
            'stop_desk': ['stop desk', 'stopdesk', 'point de relais', 'point relais', 'store desk', 'STOP DESK( si oui mettez OUI sinon laissez vide )'],
            'map': ['lien map', 'map', 'carte', 'google map', 'link', 'url']
        }
        
        # Define required columns
        REQUIRED_COLUMNS = ['name', 'phone', 'address']
        
        # Find the header row by scanning first 10 rows for keyword matches
        max_matches = 0
        header_row = 1
        
        for row in range(1, 11):  # Check first 10 rows
            matches = 0
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is not None:
                    cell_text = str(cell_value).lower().strip()
                    for _ in header_synonyms.values():
                        if any(synonym in cell_text for synonym in _):
                            matches += 1
                            break
            
            if matches > max_matches:
                max_matches = matches
                header_row = row
                
        self.header_row = header_row
        self.data_start_row = header_row + 1
        
        logger.info(f"Detected header row at row {self.header_row}, data starts at row {self.data_start_row}")
        
        # Map column indices to our field names
        # Track which columns we've already mapped to avoid duplicates
        mapped_columns = set()
        
        # Force specific column positions for wilaya code and stop desk
        self.column_mapping = {
            'wilaya_code': 5,  # Column E
            'stop_desk': 17,   # Column Q
        }
        mapped_columns.add(5)  # Column E
        mapped_columns.add(17) # Column Q
        
        # Map other columns normally
        for col in range(1, sheet.max_column + 1):
            if col in mapped_columns:
                continue
                
            cell_value = sheet.cell(row=header_row, column=col).value
            if cell_value is not None:
                cell_text = str(cell_value).lower().strip()
                
                # Try to match the column header with our field names
                for field, synonyms in header_synonyms.items():
                    if field in ['wilaya_code', 'stop_desk']:  # Skip these as they're already mapped
                        continue
                    if any(synonym in cell_text for synonym in synonyms):
                        if field not in self.column_mapping:  # Only map the first matching column
                            self.column_mapping[field] = col
                            mapped_columns.add(col)
                            break
        
        # Add one more debug logging of the complete mapping
        logger.info(f"===== FINAL COLUMN MAPPING =====")
        for field, col in self.column_mapping.items():
            column_letter = get_column_letter(col)
            header_value = sheet.cell(row=header_row, column=col).value
            logger.info(f"Field '{field}' mapped to column {column_letter} (col {col}) with header: '{header_value}'")
        logger.info(f"===== END FINAL COLUMN MAPPING =====")
        
        # Validate that we have all required columns
        missing_columns = [col for col in REQUIRED_COLUMNS if col not in self.column_mapping]
        if missing_columns:
            error_msg = f"Missing required columns in template: {', '.join(missing_columns)}"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
    def _add_order_data(self, sheet, orders):
        """
        Add order data to the sheet, preserving all template properties.
        
        Args:
            sheet: The Excel worksheet object
            orders: List of order objects
        """
        # Validate orders input
        if not orders:
            raise ValueError("No orders provided for export")
            
        if not isinstance(orders, list):
            raise ValueError(f"Expected orders to be a list, got {type(orders)}")
            
        logger.info(f"Processing {len(orders)} orders")
        
        for i, order in enumerate(orders):
            if not isinstance(order, dict):
                raise ValueError(f"Expected order to be a dictionary, got {type(order)}")
                
            if self.data_start_row is None:
                self.data_start_row = 2  # Default to row 2 if header detection failed
            row_num = self.data_start_row + i
            
            try:
                # Log the order data for debugging
                logger.info(f"Processing order {i+1}: {order.get('reference', 'NO_REF')}")
                
                # Set values using our column mapping
                self._set_cell_value(sheet, row_num, 'ref', order.get('reference', ''))
                
                # Validate customer data
                if not order.get('customer'):
                    raise ValueError(f"Missing customer data for order {order.get('reference', 'NO_REF')}")
                    
                customer = order['customer']
                if not isinstance(customer, dict):
                    raise ValueError(f"Customer data must be a dictionary, got {type(customer)}")
                
                # Required field: name
                name = customer.get('name')
                if not name:
                    raise ValueError(f"Missing required field 'name' for order {order.get('reference', 'NO_REF')}")
                self._set_cell_value(sheet, row_num, 'name', str(name))
                
                # Required field: phone
                phone = customer.get('phone')
                if not phone:
                    raise ValueError(f"Missing required field 'phone' for order {order.get('reference', 'NO_REF')}")
                self._set_cell_value(sheet, row_num, 'phone', str(phone))
                
                # Optional: phone2
                phone2 = customer.get('phone2', '')
                if phone2:
                    self._set_cell_value(sheet, row_num, 'phone2', str(phone2))
                
                # Required field: address
                address = customer.get('address')
                if not address:
                    raise ValueError(f"Missing required field 'address' for order {order.get('reference', 'NO_REF')}")
                self._set_cell_value(sheet, row_num, 'address', str(address))
                
                # Handle wilaya code
                wilaya_code = order.get('wilayaCode') or customer.get('wilayaCode') or customer.get('wilaya', '')
                if wilaya_code:
                    # Clean up wilaya code (remove any text, keep only numbers)
                    wilaya_code = ''.join(c for c in str(wilaya_code) if c.isdigit())
                    if wilaya_code:  # Only set if we have a numeric code
                        self._set_cell_value(sheet, row_num, 'wilaya_code', wilaya_code)
                        logger.info(f"Set wilaya_code to '{wilaya_code}' for order {order.get('reference', 'NO_REF')}")
                
                # Skip wilaya de livraison - keep it empty as requested
                
                # Optional: commune
                commune = order.get('commune') or customer.get('commune', '')
                if commune:
                    # Clean up commune name (remove ID part if present)
                    if '_' in str(commune):
                        commune = str(commune).split('_')[0]
                    self._set_cell_value(sheet, row_num, 'commune', str(commune))
                
                # Optional: Set other fields if they exist in the mapping
                optional_fields = {
                    'product': 'livres',  # Always set product as "livres"
                    'amount': order.get('finalAmount') or order.get('totalAmount', 0),
                    'note': order.get('notes', ''),
                    'map': customer.get('mapLink', '')
                }
                
                # Handle checkbox fields - write directly to specific columns
                checkbox_fields = {
                    'fragile': {'column': 13, 'value': order.get('fragile')},      # Column M
                    'exchange': {'column': 14, 'value': order.get('echange')},      # Column N
                    'pickup': {'column': 15, 'value': order.get('pickup')},         # Column O
                    'cod': {'column': 16, 'value': order.get('recouvrement')},      # Column P
                    'stop_desk': {'column': 17, 'value': order.get('stopDesk')}     # Column Q
                }
                
                # Process all checkbox fields
                for field, config in checkbox_fields.items():
                    value = config['value']
                    cell = sheet.cell(row=row_num, column=config['column'])
                    
                    # Log the raw value for debugging
                    logger.info(f"DEBUG: {field} raw value: {value}")
                    logger.info(f"DEBUG: {field} type: {type(value)}")

                    # Simplified check: Directly check if the value evaluates to True
                    is_checked = False
                    if isinstance(value, bool):
                        is_checked = value
                    elif isinstance(value, str):
                         is_checked = value.lower() in ['true', '1', 'oui', 'yes']
                    elif isinstance(value, int):
                        is_checked = value == 1

                    cell.value = 'OUI' if is_checked else ''
                    
                    logger.info(f"Set {field} to '{cell.value}' for order {order.get('reference', 'NO_REF')}")
                
                # Set other optional fields
                for field, value in optional_fields.items():
                    if field in self.column_mapping:
                        self._set_cell_value(sheet, row_num, field, value)
                
            except Exception as e:
                error_msg = f"Error processing row {row_num}: {str(e)}"
                logger.error(error_msg)
                self.log_errors.append(error_msg)
                raise Exception(error_msg)
    
    def _set_cell_value(self, sheet, row, field, value):
        """
        Safely set a cell value based on our column mapping.
        
        Args:
            sheet: The Excel worksheet
            row: Row number
            field: Field name in our mapping
            value: Value to set
        """
        try:
            if field not in self.column_mapping:
                return  # Skip fields we don't have a mapping for
                
            col = self.column_mapping[field]
            cell = sheet.cell(row=row, column=col)
            
            # Convert value to string and clean it
            if value is None:
                value = ''
            value = str(value).strip()
            
            # Special handling for phone numbers
            if field in ['phone', 'phone2'] and value:
                # Ensure phone numbers are treated as text
                cell.number_format = '@'
                
                # Clean up the phone number
                value = ''.join(c for c in value if c.isdigit() or c == '+')
                if not value.startswith('+'):
                    value = '+213' + value.lstrip('0')
            
            # Special handling for amounts
            elif field == 'amount' and value:
                try:
                    # Convert to float and format as currency
                    float_value = float(value)
                    cell.value = float_value
                    cell.number_format = '#,##0.00'
                    return
                except ValueError:
                    # If conversion fails, treat as text
                    pass
            
            # Set the cell value
            cell.value = value
            
        except Exception as e:
            error_msg = f"Error setting cell value for field '{field}' at row {row}: {str(e)}"
            logger.error(error_msg)
            self.log_errors.append(error_msg)
            raise Exception(error_msg)
    
    def _write_error_log(self):
        """Write all logged errors to the error log file."""
        os.makedirs(os.path.dirname(ERROR_LOG_PATH), exist_ok=True)
        with open(ERROR_LOG_PATH, 'a') as f:
            f.write(f"\n--- Error Log {datetime.now().isoformat()} ---\n")
            for error in self.log_errors:
                f.write(f"{error}\n")


def export_orders_to_ecotrack(orders, template_path=None, output_path=None):
    """
    Export orders to an EcoTrack Excel file using their template.
    
    Args:
        orders: List of order objects with customer and item details
        template_path: Path to the EcoTrack template file. If None, uses the default path.
        output_path: Path to save the Excel file. If None, a default path is used.
    
    Returns:
        Path to the saved Excel file
    """
    exporter = EcoTrackExcelExporter(template_path)
    return exporter.export_orders(orders, output_path)


# If run directly, perform test export from JSON
if __name__ == "__main__":
    import argparse
    
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Export orders to EcoTrack Excel format')
    parser.add_argument('--orders', required=True, help='Path to JSON file containing orders')
    parser.add_argument('--template', required=True, help='Path to EcoTrack Excel template')
    parser.add_argument('--output', required=True, help='Path to save the output Excel file')
    
    args = parser.parse_args()
    
    try:
        # Load orders from JSON file
        with open(args.orders, 'r', encoding='utf-8') as f:
            orders = json.load(f)
            
        # Create exporter and export
        exporter = EcoTrackExcelExporter(args.template)
        output_path = exporter.export_orders(orders, args.output)
        
        # Print the output path for the Node.js process to capture
        print(output_path)
        
    except Exception as e:
        logger.error(f"Error during export: {str(e)}")
        print(f"Error: {str(e)}", file=sys.stderr)
        sys.exit(1)